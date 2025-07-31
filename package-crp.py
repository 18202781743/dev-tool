import sys
import argparse
import requests
import json
import re
import os
import logging
from datetime import datetime
from openpyxl import load_workbook

class Colors:
    RESET = '\033[0m'
    RED = '\033[31m'      # 错误
    YELLOW = '\033[33m'   # 警告
    GREEN = '\033[32m'    # 成功信息
    GRAY = '\033[90m'     # 调试信息
    
    # 背景颜色用于状态显示
    BG_GREEN = '\033[42m'   # 绿色背景 - 成功
    BG_RED = '\033[41m'     # 红色背景 - 失败
    BG_YELLOW = '\033[43m'  # 黄色背景 - 进行中
    BG_GRAY = '\033[100m'   # 灰色背景 - 未知
    WHITE = '\033[97m'      # 白色文字，配合背景色使用

class ColoredFormatter(logging.Formatter):
    COLORS = {
        logging.ERROR: Colors.RED,
        logging.WARNING: Colors.YELLOW,
        logging.INFO: Colors.GREEN,
        logging.DEBUG: Colors.GRAY,
    }

    def format(self, record):
        # 先格式化消息
        formatted = super().format(record)
        
        # 只在终端中显示颜色，整行应用颜色
        if hasattr(sys.stdout, 'isatty') and sys.stdout.isatty():
            color = self.COLORS.get(record.levelno, '')
            if color:
                formatted = f"{color}{formatted}{Colors.RESET}"
        
        return formatted

def colorize_build_state(state):
    if not (hasattr(sys.stdout, 'isatty') and sys.stdout.isatty()):
        return state
        
    state_str = str(state).upper()
    
    # 成功状态 - 绿色背景
    if state_str in ['UPLOAD_OK', 'SUCCESS', 'OK']:
        return f"{Colors.BG_GREEN}{Colors.WHITE} {state} {Colors.RESET}"
    
    # 失败状态 - 红色背景  
    elif state_str in ['UPLOAD_GIVEUP', 'APPLY_FAILED']:
        return f"{Colors.BG_RED}{Colors.WHITE} {state} {Colors.RESET}"
    
    # 进行中状态 - 黄色背景
    elif state_str in ['APPLYING', 'UPLOADING']:
        return f"{Colors.BG_YELLOW}{Colors.WHITE} {state} {Colors.RESET}"
    
    # 未知状态 - 灰色背景
    elif state_str in ['UNKNOWN'] or not state_str:
        return f"{Colors.BG_GRAY}{Colors.WHITE} {state} {Colors.RESET}"
    
    # 其他未识别状态 - 灰色背景
    else:
        return f"{Colors.BG_GRAY}{Colors.WHITE} {state} {Colors.RESET}"

# 全局参数
class ArgsInfo:
    def __init__(self):
        # 默认值
        self.topicName = "test-xxxx" # 主题名称
        self.projectName = "xxxx" # 项目名称
        self.projectBranch = "upstream/master" # 项目分支
        self.projectTag = "5.0.0" # 自定义tag
        self.projectUpdateMode = True # 根据changelog自动更新版本号
        self.branchId = 123 # snipe分支
        self.archs = "amd64;arm64;loong64;sw64;mips64el"
        self.topicType = "test"
        self.userName = "xxxx" # crp用户名（过滤topic）
        self.token = "xxxx"
        self.verbose = False # 是否显示详细输出

        # 从配置文件读取参数
        config_path = os.path.expanduser('~/.config/dev-tool/package-crp-config.json')
        with open(config_path) as f:
            config = json.load(f)
        
        # 认证信息
        self.userId = config['auth']['userId']     # crp用户id（登陆获取token）
        self.password = config['auth']['password'] # crp用户密码

        # 从配置文件中读取其他参数（如果存在）
        if 'params' in config:
            params = config['params']
            self.topicType = params.get('topicType', self.topicType)
            self.archs = params.get('archs', self.archs)
            self.branchId = params.get('branchId', self.branchId)
            self.projectBranch = params.get('projectBranch', self.projectBranch)

argsInfo = ArgsInfo()

def setup_logging():
    level = logging.DEBUG if argsInfo.verbose else logging.INFO
    
    # 创建logger
    logger = logging.getLogger(__name__)
    logger.setLevel(level)
    
    # 避免重复添加handler
    if logger.handlers:
        logger.handlers.clear()
    
    # 创建控制台处理器
    handler = logging.StreamHandler()
    handler.setLevel(level)
    
    # 使用彩色格式化器
    formatter = ColoredFormatter(
        fmt='[%(asctime)s] [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    handler.setFormatter(formatter)
    
    logger.addHandler(handler)
    return logger


logger = setup_logging()

class ProjectInfo:
    name = "dtk6"
    id = 0
    url = "https://gitee.com/xxxxx/dtk6.git"

class TopicInfo:
    name = "test-treeland-private1"
    id = 0

class BranchInfo:
    projectId = 0
    name = "upstream/master"
    commit = "b9e8b6b"
    changelog = "chore: update changelog"

class InstanceInfo:
    Arches = "amd64"
    BaseTag = None
    Branch = "upstream/master"
    BuildID = 0
    BuildState = None
    Changelog = ["chore: update changelog"]
    Commit = "xxxx"
    History = None
    ID = 0
    ProjectID = 4305
    ProjectName = "dtkcommon-v25"
    ProjectRepoUrl = None
    SlaveNode = None
    Tag = "1"
    TagSuffix = None
    TopicID = 20825
    TopicName = "test-treeland"
    TopicType = "test"
    ChangeLogMode = True
    RepoType = "deb"
    Custom = True
    BranchID = "55"

def fetchToken():
    try:
        url = "https://crp.uniontech.com/api/login"
        headers = {
            "Content-Type": "application/json"
        }
        data = {
            "userName": argsInfo.userId,
            "password": argsInfo.password
        }

        response = requests.post(
            url,
            headers=headers,
            data=json.dumps(data),
            timeout=30
        )
        response.raise_for_status()  # Raises HTTPError for bad responses

        result = response.json()
        token = result.get("Token", "")
        if not token:
            logger.error("Token not found in response")
            return ""

        return token

    except requests.exceptions.RequestException as e:
        logger.error(f"Login request failed: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Server response: {e.response.text}")
        return ""
    except json.JSONDecodeError as e:
        logger.error(f"Failed to decode response JSON: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Raw response: {e.response.text}")
        return ""

def fetchUser():
    try:
        url = "https://crp.uniontech.com/api/user"
        headers = {
            "Authorization": f"Bearer {argsInfo.token}"
        }

        response = requests.get(
            url,
            headers=headers,
            timeout=30
        )
        response.raise_for_status()

        result = response.json()
        name = result.get("Name", "")
        if not name:
            logger.error("Name not found in response")
            return ""

        return name

    except requests.exceptions.RequestException as e:
        logger.error(f"Fetch user request failed: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Server response: {e.response.text}")
        return ""
    except json.JSONDecodeError as e:
        logger.error(f"Failed to decode response JSON: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Raw response: {e.response.text}")
        return ""

def listPojects():
    try:
        url = "https://crp.uniontech.com/api/project"
        headers = {
            "Authorization": f"Bearer {argsInfo.token}",
            "Content-Type": "application/json"
        }
        data = {
            "page": 0,
            "perPage": 0,
            "projectGroupID": 0,
            "newCommit": False,
            "archived": False,
            "branchID": argsInfo.branchId,
            "name": argsInfo.projectName
        }

        response = requests.post(
            url,
            headers=headers,
            json=data,
            timeout=30
        )
        response.raise_for_status()

        projects = []
        result = response.json()
        projects_data = result.get("Projects", [])
        
        for project in projects_data:
            projectName = project.get("Name", "")
            id = project.get("ID", 0)
            repoUrl = project.get("RepoUrl", "")
            
            if re.search(argsInfo.projectName, projectName, re.IGNORECASE):
                info = ProjectInfo()
                info.id = id
                info.name = projectName
                info.url = repoUrl
                projects.append(info)

        return projects

    except requests.exceptions.RequestException as e:
        logger.error(f"List projects request failed: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Server response: {e.response.text}")
        return []
    except json.JSONDecodeError as e:
        logger.error(f"Failed to decode projects response: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Raw response: {e.response.text}")
        return []
    except Exception as e:
        logger.error(f"Unexpected error in listPojects: {str(e)}")
        return []

def listTopics():
    try:
        url = "https://crp.uniontech.com/api/topics/search"
        headers = {
            "Authorization": f"Bearer {argsInfo.token}",
            "Content-Type": "application/json"
        }
        data = {
            "TopicType": argsInfo.topicType,
            "UserName": argsInfo.userName,
            "BranchID": argsInfo.branchId
        }

        response = requests.post(
            url,
            headers=headers,
            json=data,
            timeout=30
        )
        response.raise_for_status()

        topics = []
        result = response.json()
        
        for topic in result:
            id = topic.get("ID", 0)
            topicName = topic.get("Name", "")
            
            if re.search(argsInfo.topicName, topicName, re.IGNORECASE):
                info = TopicInfo()
                info.id = id
                info.name = topicName
                topics.append(info)

        return topics

    except requests.exceptions.RequestException as e:
        logger.error(f"List topics request failed: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Server response: {e.response.text}")
        return []
    except json.JSONDecodeError as e:
        logger.error(f"Failed to decode topics response: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Raw response: {e.response.text}")
        return []
    except Exception as e:
        logger.error(f"Unexpected error in listTopics: {str(e)}")
        return []

def fetchCommitInfo(repoUrl, commit):
    try:
        url = "https://crp.uniontech.com/api/projects/getGerritCommitMessage"
        headers = {
            "Authorization": f"Bearer {argsInfo.token}",
            "Content-Type": "application/json"
        }
        data = {
            "repo_url": repoUrl,
            "commit_id": commit
        }

        response = requests.post(
            url,
            headers=headers,
            json=data,
            timeout=30
        )
        response.raise_for_status()

        if response.status_code != 200:
            logger.error(f"Fetch commit info failed: {response.status_code}")
            return ""

        result = response.json()
        message = result.get("message", "")
        if not message:
            logger.warning("No commit message found in response")
        return message

    except requests.exceptions.RequestException as e:
        logger.error(f"Fetch commit info failed: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Server response: {e.response.text}")
        return ""
    except json.JSONDecodeError as e:
        logger.error(f"Failed to decode commit info response: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Raw response: {e.response.text}")
        return ""
    except Exception as e:
        logger.error(f"Unexpected error in fetchCommitInfo: {str(e)}")
        return ""

def listBranchs(projectId, projectUrl, targetName):
    try:
        url = f"https://crp.uniontech.com/api/projects/{projectId}/branches"
        headers = {
            "Authorization": f"Bearer {argsInfo.token}"
        }

        response = requests.get(
            url,
            headers=headers,
            timeout=30
        )
        response.raise_for_status()

        branchs = []
        result = response.json()
        
        for branch in result:
            commit = branch.get("Commit", "")
            name = branch.get("Name", "")
            
            if re.search(targetName, name, re.IGNORECASE):
                info = BranchInfo()
                info.commit = commit
                info.name = name
                info.projectId = projectId
                info.changelog = fetchCommitInfo(projectUrl, commit) == "" and branch.get("Message", "") or fetchCommitInfo(projectUrl, commit)
                branchs.append(info)

        return branchs

    except requests.exceptions.RequestException as e:
        logger.error(f"List branches failed: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Server response: {e.response.text}")
        return []
    except json.JSONDecodeError as e:
        logger.error(f"Failed to decode branches response: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Raw response: {e.response.text}")
        return []
    except Exception as e:
        logger.error(f"Unexpected error in listBranchs: {str(e)}")
        return []

def listCreatedInstances(topicId):
    try:
        url = f"https://crp.uniontech.com/api/topics/{topicId}/releases"
        headers = {
            "Authorization": f"Bearer {argsInfo.token}"
        }

        response = requests.get(
            url,
            headers=headers,
            timeout=30
        )
        response.raise_for_status()

        instances = []
        result = response.json()
        
        for instance in result:
            info = InstanceInfo()
            info.ID = instance.get("ID", 0)
            info.ProjectID = instance.get("ProjectID", 0)
            info.ProjectName = instance.get("ProjectName", "")
            info.Branch = instance.get("Branch", "")
            info.Tag = instance.get("Tag", "")
            build_state = instance.get("BuildState", {})
            info.BuildState = build_state.get("state", "unknown")
            info.BuildID = instance.get("BuildID", 0)
            instances.append(info)

        return instances

    except requests.exceptions.RequestException as e:
        logger.error(f"List instances failed: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Server response: {e.response.text}")
        return []
    except json.JSONDecodeError as e:
        logger.error(f"Failed to decode instances response: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Raw response: {e.response.text}")
        return []
    except Exception as e:
        logger.error(f"Unexpected error in listCreatedInstances: {str(e)}")
        return []

def deleteInstance(instanceId):
    try:
        url = f"https://crp.uniontech.com/api/topic_releases/{instanceId}"
        headers = {
            "Authorization": f"Bearer {argsInfo.token}"
        }

        response = requests.delete(
            url,
            headers=headers,
            timeout=30
        )
        response.raise_for_status()
        logger.info(f"Successfully deleted instance: {instanceId}")

    except requests.exceptions.RequestException as e:
        logger.error(f"Delete instance {instanceId} failed: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Server response: {e.response.text}")
    except Exception as e:
        logger.error(f"Unexpected error in deleteInstance: {str(e)}")

def createInstance(instanceInfo):
    try:
        url = f"https://crp.uniontech.com/api/topics/{instanceInfo.TopicID}/new_release"
        headers = {
            "Authorization": f"Bearer {argsInfo.token}",
            "Content-Type": "application/json"
        }
        data = {
            "Arches": instanceInfo.Arches,
            "BaseTag": instanceInfo.BaseTag,
            "Branch": instanceInfo.Branch,
            "BuildID": instanceInfo.BuildID,
            "BuildState": instanceInfo.BuildState,
            "Changelog": [instanceInfo.Changelog],
            "Commit": instanceInfo.Commit,
            "History": instanceInfo.History,
            "ID": instanceInfo.ID,
            "ProjectID": instanceInfo.ProjectID,
            "ProjectName": instanceInfo.ProjectName,
            "ProjectRepoUrl": instanceInfo.ProjectRepoUrl,
            "SlaveNode": instanceInfo.SlaveNode,
            "Tag": instanceInfo.Tag,
            "TagSuffix": instanceInfo.TagSuffix,
            "TopicID": instanceInfo.TopicID,
            "TopicType": instanceInfo.TopicType,
            "ChangeLogMode": instanceInfo.ChangeLogMode,
            "RepoType": instanceInfo.RepoType,
            "Custom": instanceInfo.Custom,
            "BranchID": instanceInfo.BranchID
        }

        response = requests.post(
            url,
            headers=headers,
            json=data,
            timeout=30
        )
        response.raise_for_status()
        logger.info(f"Successfully created instance: {response.text}")
        logger.debug(f"Instance creation response: {response.json()}")

    except requests.exceptions.RequestException as e:
        logger.error(f"Create instance failed: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Server response: {e.response.text}")
        logger.debug(f"Request data: {data}")
        raise
    except json.JSONDecodeError as e:
        logger.error(f"Failed to decode create instance response: {str(e)}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"Raw response: {e.response.text}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error in createInstance: {str(e)}")
        raise

def listInstances():
    instances = []
    topics = listTopics()
    if len(topics) == 0:
        logger.warning("No topics found matching criteria")
        return []

    for topic in topics:
        projects = listPojects()
        if len(projects) == 0:
            continue
        for project in projects:
            branchs = listBranchs(project.id, project.url, argsInfo.projectBranch)
            if len(branchs) == 0:
                continue
            for branch in branchs:
                info = InstanceInfo()
                info.Commit = branch.commit
                info.Branch = branch.name
                info.Arches = argsInfo.archs
                info.BranchID = argsInfo.branchId
                info.TopicType = argsInfo.topicType
                info.TopicID = topic.id
                info.TopicName = topic.name
                info.ProjectID = project.id
                info.ProjectName = project.name
                info.Changelog = branch.changelog
                info.Tag = argsInfo.projectTag
                info.ChangeLogMode = argsInfo.projectUpdateMode
                instances.append(info)

    return instances

def createOrUpdate():
    instances = listInstances()
    for item in instances:
        logger.info(f"Creating instance - Topic: {item.TopicName}, Project: {item.ProjectName}, Branch: {item.Branch}, Changelog: {item.Changelog}")
        createdInstance = listCreatedInstances(item.TopicID)
        for (createdInstance) in createdInstance:
            if (createdInstance.ProjectName == item.ProjectName and createdInstance.Branch == item.Branch):
                deleteInstance(createdInstance.ID)
        createInstance(item)

def generate_table(topic_name, template_path=None):
    # 1. 优先使用参数指定的模板路径
    if template_path and os.path.exists(template_path):
        tpl_path = template_path
    else:
        # 2. 当前目录查找
        local_tpl = os.path.join(os.getcwd(), "crp-gendoc.xlsx")
        config_tpl = os.path.expanduser("~/.config/dev-tool/crp-gendoc.xlsx")
        if os.path.exists(local_tpl):
            tpl_path = local_tpl
        elif os.path.exists(config_tpl):
            tpl_path = config_tpl
        else:
            logger.error("未找到模板文件crp-gendoc.xlsx，请在当前目录或~/.config/dev-tool/下放置模板，或用--template指定路径！")
            return
    topics = listTopics()
    topic_id = None
    for topic in topics:
        if topic.name == topic_name:
            topic_id = topic.id
            break
    if not topic_id:
        logger.error(f"未找到主题: {topic_name}")
        return
    releases_url = f"https://crp.uniontech.com/api/topics/{topic_id}/releases"
    headers = {"Authorization": f"Bearer {argsInfo.token}"}
    try:
        response = requests.get(releases_url, headers=headers, timeout=30)
        response.raise_for_status()
        releases = response.json()
    except Exception as e:
        logger.error(f"获取releases失败: {e}")
        return
    if not releases:
        logger.info("无release数据，不插入新行")
        return
    wb = load_workbook(tpl_path)
    if "ChangeLog" not in wb.sheetnames:
        logger.error("模板文件中缺少ChangeLog工作表")
        return
    ws = wb["ChangeLog"]
    module_name_row = None
    test_desc_row = None
    style_template_row = None
    pack_branch_row = None
    for row_idx in range(1, 100):
        cell_value = ws.cell(row=row_idx, column=1).value
        if cell_value and "模块名" in str(cell_value):
            module_name_row = row_idx
            style_template_row = row_idx + 1
        elif cell_value == "转测说明":
            test_desc_row = row_idx
        elif cell_value and "打包分支及测试主题" in str(cell_value):
            pack_branch_row = row_idx
    if not module_name_row or not test_desc_row or not style_template_row:
        logger.error("未找到关键行位置")
        return
    insert_row = module_name_row + 1
    if insert_row >= test_desc_row:
        logger.error("位置异常：模块名行在转测说明行之后")
        return
    max_columns = ws.max_column
    reserved_rows = test_desc_row - insert_row
    if len(releases) > reserved_rows:
        logger.error(f"数据行数({len(releases)})超过模板预留空白行数({reserved_rows})，请扩展模板！")
        return
    for idx, release in enumerate(releases):
        row_idx = insert_row + idx
        module_name = release.get("SourcePkgName", "")
        tag = release.get("Tag", "")
        commit = release.get("Commit", "")
        row_values = [module_name, module_name, tag, commit] + [None] * (max_columns - 4)
        for col in range(1, max_columns + 1):
            dest_cell = ws.cell(row=row_idx, column=col)
            dest_cell.value = row_values[col-1]
    # 填写repoUrls到"打包分支及测试主题*"行的下一列（B列）
    if pack_branch_row:
        repo_urls = []
        # 获取topic_urls，确保在所有使用前定义
        topic_urls_url = "https://crp.uniontech.com/api/topic_urls"
        try:
            data = {
                "branchID": argsInfo.branchId,
                "topicID": topic_id
            }
            response = requests.post(topic_urls_url, headers=headers, json=data, timeout=30)
            response.raise_for_status()
            topic_urls = response.json()
        except Exception as e:
            logger.error(f"获取topic_urls失败: {e}")
            topic_urls = []
        if isinstance(topic_urls, dict):
            repo_urls_list = topic_urls.get("RepoUrls", [])
            if isinstance(repo_urls_list, list):
                repo_urls.extend(repo_urls_list)
            elif isinstance(repo_urls_list, str) and repo_urls_list:
                repo_urls.append(repo_urls_list)
        if not repo_urls:
            logger.error("未获取到任何有效的RepoUrls，请检查topic_urls接口返回！")
            return
        repo_urls_str = ", ".join(repo_urls)
        ws.cell(row=pack_branch_row, column=2, value=repo_urls_str)
    filename = f"测试-桌面专业版-转测申请单-{topic_name}.xlsx"
    wb.save(filename)
    logger.info(f"表格已生成: {filename}")

def main(argv):
    parser = argparse.ArgumentParser(description='Pack for CRP.')
    parser.add_argument('command', nargs='?', default='pack', choices=['pack', 'test', 'projects', 'topics', 'instances', 'branches', 'gendoc'], help='The command type (list or pack)')

    parser.add_argument('--topic', type=str, default=None, help='The topic name parameter')
    parser.add_argument('--name', type=str, default=None, help='The project name parameter')
    parser.add_argument('--branch', type=str, default=None, help='The project branch parameter')
    parser.add_argument('--tag', type=str, default=None, help='The project tag parameter')
    parser.add_argument('--verbose', action='store_true', help='Show verbose debug output')
    parser.add_argument('--template', type=str, default=None, help='Template file path (optional)')

    args = parser.parse_args()
    argsInfo.verbose = args.verbose
    global logger
    logger = setup_logging()  # Reinitialize logger with new level

    if (args.topic is not None):
        argsInfo.topicName = args.topic
    if (args.name is not None):
        argsInfo.projectName = args.name
    if (args.branch is not None):
        argsInfo.projectBranch = args.branch
    if (args.tag is not None):
        argsInfo.projectTag = args.tag
        argsInfo.projectUpdateMode = False
    
    token = fetchToken()
    argsInfo.token = token
    userName = fetchUser()
    argsInfo.userName = userName
    
    if (args.command == 'projects'):
        projects = listPojects()
        for project in projects:
            logger.info(f"Found project: {project.name}")
    if (args.command == 'topics'):
        topics = listTopics()
        for topic in topics:
            logger.info(f"Found topic: {topic.name}")
    if (args.command == 'instances'):
        topics = listTopics()
        if (len(topics) == 0):
            logger.warning("No topics found")
            return
        for topic in topics:
            instances = listCreatedInstances(topic.id)
            for instance in instances:
                colored_state = colorize_build_state(instance.BuildState)
                if instance.BuildState == "UPLOAD_OK":
                    logger.info(f"Instance found - Topic: {topic.name}, Project: {instance.ProjectName}, Branch: {instance.Branch}, Tag: {instance.Tag}, State: {colored_state}")
                else:
                    logger.info(f"Instance found - Topic: {topic.name}, Project: {instance.ProjectName}, Branch: {instance.Branch}, Tag: {instance.Tag}, State: {colored_state}\n查看详情: https://shuttle.uniontech.com/#/tasks/task?taskid={instance.BuildID}")
    if (args.command == 'test'):
        instances = listInstances()
        for item in instances:
            logger.info(f"Test instance - Topic: {item.TopicName}, Project: {item.ProjectName}, Branch: {item.Branch}, Changelog: {item.Changelog}")
    if (args.command == 'branches'):
        topics = listTopics()
        if len(topics) == 0:
            logger.warning("No topics found for branches listing")
            return []
        for topic in topics:
            projects = listPojects()
            if len(projects) == 0:
                continue
            for project in projects:
                branchs = listBranchs(project.id, project.url, "")
                for branch in branchs:
                    logger.info(f"Branch info - Topic: {topic.name}, Project: {project.name}, Branch: {branch.name}, Changelog: {branch.changelog}")
    if (args.command == 'gendoc'):
        if not args.topic:
            logger.error("必须指定--topic参数")
            return
        generate_table(args.topic, template_path=args.template)
    if (args.command == 'pack'):
        createOrUpdate()

if(__name__=="__main__"):
    os.environ.pop("https_proxy", None)
    os.environ.pop("http_proxy", None)
    os.environ.pop("all_proxy", None)
    main(sys.argv)
